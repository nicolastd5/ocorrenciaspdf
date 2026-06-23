"""
Módulo de processamento de ocorrências.
Extrai dados do PDF e atualiza a planilha Excel.
"""

import logging
import shutil
import os
import re

import pdfplumber
from openpyxl import load_workbook

logger = logging.getLogger("processador")


class ProcessadorOcorrencias:
    """Processador principal de ocorrências PDF → Excel."""

    TODOS_CODIGOS = ['FA', 'AT', 'A-', 'SD', 'LC', 'AA', 'AP', 'LM', 'FE', '14', '13']
    SEM_QUANTIDADE = ['AP', 'LM', 'FE']
    ORDEM = ['FA', 'AT', 'A-', 'SD', 'LC', 'AA', 'AP', 'LM', 'FE', '14', '13']
    CODIGOS_DEDUZIR = ['FA', 'AT', 'SD', 'LC']
    COLUNAS_QT = ['qt va', 'qt vr', 'qt vt']
    VU_VT_HEADER = 'vu vt'
    DESCRICOES = {
        'FA': 'Faltas',
        'AT': 'Atestado',
        'A-': 'Declaração Horas Negativas',
        'SD': 'Suspensão Disciplinar',
        'LC': 'Licença Casamento',
        'AA': 'Ausência Autorizada',
        'AP': 'Afastamento Previdenciário',
        'LM': 'Afastamento Maternidade',
        'FE': 'Férias',
        '14': 'Luto',
        '13': 'Falecimento',
    }

    def extrair_ocorrencias(self, pdf_path, codigos_alvo):
        """
        Extrai ocorrências do PDF de jornada de trabalho.

        Args:
            pdf_path: Caminho do arquivo PDF.
            codigos_alvo: Set de códigos a procurar.

        Returns:
            dict: {codigo_re: {'nome': str, 'ocorrencias': {codigo: contagem}}}
        """
        resultados = {}
        codigos_set = set(codigos_alvo)

        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                tabelas = pagina.extract_tables()
                if not tabelas:
                    continue

                for tabela in tabelas:
                    for linha in tabela:
                        if not linha or len(linha) < 2 or not linha[0] or not linha[1]:
                            continue

                        nome = str(linha[0]).strip() if linha[0] is not None else ''
                        codigo = str(linha[1]).strip() if linha[1] is not None else ''

                        if codigo == 'Código' or not codigo.isdigit():
                            continue

                        ocorrencias = {}
                        for celula in linha[6:]:
                            if celula is None:
                                continue
                            cod = str(celula).strip()
                            if cod in codigos_set:
                                ocorrencias[cod] = ocorrencias.get(cod, 0) + 1

                        if ocorrencias:
                            if codigo not in resultados:
                                resultados[codigo] = {'nome': nome, 'ocorrencias': {}}
                            for k, v in ocorrencias.items():
                                resultados[codigo]['ocorrencias'][k] = \
                                    resultados[codigo]['ocorrencias'].get(k, 0) + v

        return resultados

    def extrair_ocorrencias_texto(self, pdf_path, codigos_alvo):
        """
        Segunda varredura: extrai ocorrências via extract_text() + regex posicional.
        Independente de detecção de tabelas. Mesmo formato de retorno que
        extrair_ocorrencias: {re: {'nome': str, 'ocorrencias': {codigo: contagem}}}
        """
        resultados = {}
        codigos_set = set(codigos_alvo)
        # RE: 5+ dígitos no início ou após espaços, precedido de nome
        re_linha = re.compile(r'^(.+?)\s{2,}(\d{5,})\b')

        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if not texto:
                    continue
                for linha in texto.splitlines():
                    m = re_linha.match(linha.strip())
                    if not m:
                        continue
                    nome = m.group(1).strip()
                    codigo_re = m.group(2).strip()
                    tokens = linha.split()
                    ocorrencias = {}
                    for tok in tokens:
                        tok_clean = tok.strip()
                        if tok_clean in codigos_set:
                            ocorrencias[tok_clean] = ocorrencias.get(tok_clean, 0) + 1
                    if ocorrencias:
                        if codigo_re not in resultados:
                            resultados[codigo_re] = {'nome': nome, 'ocorrencias': {}}
                        for k, v in ocorrencias.items():
                            resultados[codigo_re]['ocorrencias'][k] = (
                                resultados[codigo_re]['ocorrencias'].get(k, 0) + v
                            )
        return resultados

    def reconciliar(self, resultados, codigos_alvo):
        """
        Compara resultados de 2 ou 3 camadas de extração.

        Args:
            resultados: lista de dicts no formato {re: {'nome', 'ocorrencias'}}
                        Ordem: [v1, v2] ou [v1, v2, ia]
            codigos_alvo: lista de códigos a considerar

        Returns:
            {
              'concordantes': {re: {'nome', 'ocorrencias'}},
              'conflitos': [{re, nome, codigo, valores, sugestao}]
            }
        """
        from collections import Counter

        nomes = ['v1', 'v2', 'ia']
        camadas = resultados

        todos_res = set()
        for c in camadas:
            todos_res.update(c.keys())

        concordantes = {}
        conflitos = []

        for re_val in todos_res:
            nome = next(
                (c[re_val]['nome'] for c in camadas if re_val in c), ''
            )

            todos_codigos = set()
            for c in camadas:
                if re_val in c:
                    todos_codigos.update(c[re_val]['ocorrencias'].keys())
            todos_codigos = todos_codigos.intersection(set(codigos_alvo))

            re_conflitos = []
            ocorrencias_finais = {}

            for cod in todos_codigos:
                valores_por_camada = {}
                for i, c in enumerate(camadas):
                    chave = nomes[i]
                    val = c.get(re_val, {}).get('ocorrencias', {}).get(cod, 0)
                    valores_por_camada[chave] = val

                vals = list(valores_por_camada.values())
                counter = Counter(vals)
                valor_majoritario, votos = counter.most_common(1)[0]

                todos_iguais = len(set(vals)) == 1
                maioria_clara = votos > len(camadas) / 2

                if todos_iguais or maioria_clara:
                    ocorrencias_finais[cod] = valor_majoritario
                else:
                    sugestao = max(vals)
                    re_conflitos.append({
                        're': re_val,
                        'nome': nome,
                        'codigo': cod,
                        'valores': valores_por_camada,
                        'sugestao': sugestao,
                    })

            if re_conflitos:
                conflitos.extend(re_conflitos)
            # Os códigos em que as camadas concordaram são mantidos mesmo quando
            # o RE tem outros códigos em conflito — a resolução do usuário só
            # complementa os conflitantes.
            if ocorrencias_finais:
                concordantes[re_val] = {'nome': nome, 'ocorrencias': ocorrencias_finais}

        return {'concordantes': concordantes, 'conflitos': conflitos}

    def verificar_com_ia(self, pdf_path, codigos_alvo, api_key, modelo):
        """
        Terceira camada opcional: Gemini Vision re-extrai ocorrências a partir
        de imagens das páginas do PDF.

        Args:
            pdf_path: Caminho do arquivo PDF.
            codigos_alvo: Lista de códigos a procurar.
            api_key: API Key do Google Gemini.
            modelo: Nome do modelo Gemini a usar.

        Returns:
            dict no formato {re: {'nome', 'ocorrencias'}} ou None em caso de erro.
        """
        if not api_key:
            return None

        try:
            import pypdfium2 as pdfium
            import google.genai as genai
            import json as _json
            import time as _time

            client = genai.Client(api_key=api_key)

            codigos_str = ', '.join(codigos_alvo)
            prompt = (
                f"Analise esta folha de ponto. Para cada linha que contenha um RE numérico "
                f"(número de matrícula com 5+ dígitos), identifique o RE, o nome do funcionário "
                f"e a contagem de cada código de ocorrência presente na linha. "
                f"Códigos a procurar: {codigos_str}. "
                f"Responda APENAS em JSON válido, sem markdown, no formato: "
                f'[{{"re": "12345", "nome": "NOME", "ocorrencias": {{"AT": 2, "FA": 1}}}}]'
            )

            doc = pdfium.PdfDocument(pdf_path)
            resultados = {}

            for i in range(len(doc)):
                if i > 0:
                    _time.sleep(4)
                page = doc[i]
                bitmap = page.render(scale=2)
                img = bitmap.to_pil()

                response = client.models.generate_content(
                    model=modelo,
                    contents=[prompt, img],
                )
                raw = response.text.strip()

                if raw.startswith('```'):
                    raw = raw.split('```')[1]
                    if raw.startswith('json'):
                        raw = raw[4:]
                    raw = raw.strip()

                try:
                    registros = _json.loads(raw)
                except _json.JSONDecodeError:
                    continue

                for reg in registros:
                    re_val = str(reg.get('re', '')).strip()
                    nome = str(reg.get('nome', '')).strip()
                    ocorr = reg.get('ocorrencias', {})
                    if not re_val:
                        continue
                    if re_val not in resultados:
                        resultados[re_val] = {'nome': nome, 'ocorrencias': {}}
                    for cod, cnt in ocorr.items():
                        if cod in set(codigos_alvo):
                            resultados[re_val]['ocorrencias'][cod] = (
                                resultados[re_val]['ocorrencias'].get(cod, 0) + int(cnt)
                            )

            return resultados if resultados else {}

        except Exception:
            # None sinaliza fallback para V1+V2; o log preserva o diagnóstico
            # (timeout, quota 429, modelo inexistente...) que antes se perdia.
            logger.warning("verificar_com_ia falhou — usando fallback", exc_info=True)
            return None

    def montar_motivo(self, ocorrencias, codigos_selecionados):
        """
        Monta a string de motivo a partir das ocorrências.

        Regras:
        - Ordem: FA, AT, SD, LC, AA, AP, LM
        - Quantidade na frente quando > 1 (ex: 2 AT, 3 FA)
        - AP e LM nunca recebem quantidade
        - Múltiplos códigos separados por vírgula
        """
        partes = []
        codigos_set = set(codigos_selecionados)

        for codigo in self.ORDEM:
            if codigo in ocorrencias and codigo in codigos_set:
                contagem = ocorrencias[codigo]
                if codigo in self.SEM_QUANTIDADE:
                    partes.append(codigo)
                else:
                    if contagem > 1:
                        partes.append(f"{contagem} {codigo}")
                    else:
                        partes.append(codigo)

        return ', '.join(partes)

    def processar(self, pdf_path, xlsx_path, output_path, codigos,
                  progress_cb=None, dias_mes=None, colunas_qt_sel=None,
                  dados_externos=None):
        """
        Processa os arquivos e retorna um dicionário com os resultados.

        Args:
            pdf_path: Caminho do PDF de faltas.
            xlsx_path: Caminho da planilha Excel de pedido.
            output_path: Caminho para salvar a planilha atualizada.
            codigos: Lista de códigos a incluir.
            progress_cb: Callable(pct: int, msg: str) para atualizar progresso.

        Returns:
            dict: {
                'total_pdf': int,
                'matched': int,
                'atualizados': [{'re': str, 'nome': str, 'motivo': str}],
                'nao_encontrados': [{'re': str, 'nome': str, 'motivo': str}]
            }
        """
        def _prog(pct, msg):
            if progress_cb:
                progress_cb(pct, msg)

        # 1. Extrair ocorrências do PDF
        _prog(5, "Lendo PDF...")
        if dados_externos is not None:
            resultados_pdf = dados_externos
            _prog(50, "Dados reconciliados recebidos. Abrindo planilha...")
        else:
            resultados_pdf = self.extrair_ocorrencias(pdf_path, codigos)
            _prog(50, "PDF lido. Abrindo planilha...")

        # 2. Copiar e abrir planilha
        shutil.copy(xlsx_path, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        _prog(60, "Planilha aberta. Cruzando dados...")

        # 3. Encontrar colunas RE, MOTIVO, Qt VA/VR/VT e Vu VT
        re_col = None
        motivo_col = None
        qt_cols = {}  # {'qt va': col_num, ...}
        vu_vt_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                val_lower = str(val).lower().strip()
                if motivo_col is None and val_lower == 'motivo':
                    motivo_col = col
                if re_col is None and val_lower == 'folha re':
                    re_col = col
                if val_lower in self.COLUNAS_QT:
                    qt_cols[val_lower] = col
                if val_lower == self.VU_VT_HEADER:
                    vu_vt_col = col

        if not re_col or not motivo_col:
            raise ValueError(
                f"Colunas não encontradas na planilha. "
                f"RE col: {re_col}, MOTIVO col: {motivo_col}. "
                f"Verifique se a planilha tem colunas de matrícula/RE e MOTIVO."
            )

        # 4. Cruzar dados
        excel_res = set()
        matched = 0
        atualizados = []

        total_rows = ws.max_row - 1
        for i, row in enumerate(range(2, ws.max_row + 1)):
            re_val = ws.cell(row=row, column=re_col).value
            if re_val is not None:
                re_str = str(int(re_val)) if isinstance(re_val, (float, int)) else str(re_val).strip()
                excel_res.add(re_str)

                # Colunas Qt ativas = interseção entre detectadas na planilha e selecionadas pelo usuário
                qt_cols_ativas = {
                    k: v for k, v in qt_cols.items()
                    if colunas_qt_sel is None or k in colunas_qt_sel
                }

                if dias_mes is not None and qt_cols_ativas:
                    for col_nome, qt_col in qt_cols_ativas.items():
                        if col_nome == 'qt vt' and vu_vt_col is not None:
                            if not ws.cell(row=row, column=vu_vt_col).value:
                                continue
                        ws.cell(row=row, column=qt_col).value = dias_mes

                if re_str in resultados_pdf:
                    ocorr = resultados_pdf[re_str]['ocorrencias']
                    motivo = self.montar_motivo(ocorr, codigos)
                    if motivo:
                        ws.cell(row=row, column=motivo_col).value = motivo
                        matched += 1
                        atualizados.append({
                            're': re_str,
                            'nome': resultados_pdf[re_str]['nome'],
                            'motivo': motivo
                        })

                    if dias_mes is not None and qt_cols_ativas:
                        dias_ded = sum(ocorr.get(c, 0) for c in self.CODIGOS_DEDUZIR)
                        if dias_ded > 0:
                            for col_nome, qt_col in qt_cols_ativas.items():
                                # Qt VT: só deduz se a célula Vu VT do RE tiver valor
                                if col_nome == 'qt vt' and vu_vt_col is not None:
                                    vu_vt_val = ws.cell(row=row, column=vu_vt_col).value
                                    if not vu_vt_val:
                                        continue
                                ws.cell(row=row, column=qt_col).value = max(0, dias_mes - dias_ded)
            if total_rows > 0:
                pct = 60 + int((i / total_rows) * 30)
                _prog(pct, f"Cruzando dados... ({i}/{total_rows})")

        # 5. Encontrar não correspondidos
        _prog(90, "Finalizando...")
        nao_encontrados = []
        for codigo, dados in resultados_pdf.items():
            motivo = self.montar_motivo(dados['ocorrencias'], codigos)
            if motivo and codigo not in excel_res:
                nao_encontrados.append({
                    're': codigo,
                    'nome': dados['nome'],
                    'motivo': motivo
                })

        nao_encontrados.sort(key=lambda x: x['nome'])

        # 5b. Gravar aba "Não localizados" na própria planilha de saída.
        if 'Não localizados' in wb.sheetnames:
            del wb['Não localizados']
        if nao_encontrados:
            ws_nl = wb.create_sheet('Não localizados')
            ws_nl.append(['Folha RE', 'Nome', 'Motivo'])
            for item in nao_encontrados:
                ws_nl.append([item['re'], item['nome'], item['motivo']])

        # 6. Salvar
        _prog(97, "Salvando planilha...")
        wb.save(output_path)
        _prog(100, "Concluído!")

        return {
            'total_pdf': len(resultados_pdf),
            'matched': matched,
            'atualizados': atualizados,
            'nao_encontrados': nao_encontrados,
        }
