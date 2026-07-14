import csv
import os
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook
import pdfplumber
import xlrd


# ─── Adapter xlsx → API do xlrd ──────────────────────────────────────────────
# Permite que o codigo legado (que usa ws.cell_value/nrows/ncols e wb.datemode)
# leia tambem arquivos .xlsx sem reescrever as funcoes existentes.

class _XlsxSheetAdapter:
    """Imita a API minima de uma sheet xlrd a partir de uma worksheet openpyxl."""

    def __init__(self, ws):
        self.name = ws.title
        self._rows = []
        for row in ws.iter_rows(values_only=True):
            self._rows.append(list(row))
        self.nrows = len(self._rows)
        self.ncols = max((len(r) for r in self._rows), default=0)

    def cell_value(self, row, col):
        if row >= self.nrows:
            return ''
        linha = self._rows[row]
        if col >= len(linha):
            return ''
        v = linha[col]
        if v is None:
            return ''
        if isinstance(v, datetime):
            # Devolve string ISO; _formatar_data reconhece esse formato.
            return v.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(v, date):
            return v.strftime('%Y-%m-%d')
        return v


class _XlsxBookAdapter:
    """Imita wb xlrd para arquivos .xlsx."""

    def __init__(self, xlsx_path):
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        self._sheets = [_XlsxSheetAdapter(ws) for ws in wb.worksheets]
        self.nsheets = len(self._sheets)
        # No xlrd datemode 0 = 1900, 1 = 1904. openpyxl: epoch 1904 = data 1904-01-01.
        try:
            self.datemode = 1 if wb.epoch.year == 1904 else 0
        except Exception:
            self.datemode = 0
        wb.close()

    def sheet_by_index(self, idx):
        return self._sheets[idx]


def _abrir_workbook_cadastral(path):
    """Abre cadastral xls/xlsx e devolve objeto com a API esperada pelo codigo legado."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xlsx':
        return _XlsxBookAdapter(path)
    return xlrd.open_workbook(path)
# ──────────────────────────────────────────────────────────────────────────────

COLUNAS_CSV = [
    'CNPJ', 'CEP', 'LOGRADOURO', 'NÚMERO', 'COMPLEMENTO', 'PONTO REFERENCIA',
    'UF', 'ESTADO', 'MATRÍCULA', 'NOME DO FUNCIONÁRIO', 'CPF', 'RG',
    'DATA DE NASCIMENTO', 'CARGO', 'DEPARTAMENTO', 'NOME DA MÃE',
    'BENEFÍCIO DO FUNCIONÁRIO', 'VALOR UNITÁRIO', 'QUANTIDADE DIÁRIA',  # Coluna Q
    'PERÍODO DE DIAS TRABALHADOS',  # Coluna T
    'TIPO VALOR',   # Coluna U
    'REDE RECARGA', # Coluna V — valor 3 quando operadora for RIOCARD
]

# Endereço fixo da empresa (primeira parte do CSV — linhas antes da MATRÍCULA)
EMPRESA_CNPJ         = '02.738.552/0001-27'
EMPRESA_CEP          = '06455000'
EMPRESA_LOGRADOURO   = 'ALAMEDA ARAGUAIA'
EMPRESA_NUMERO       = '3354'
EMPRESA_COMPLEMENTO  = 'ALPHAVILLE INDUSTRIAL'
EMPRESA_UF           = 'SP'
EMPRESA_ESTADO       = 'BARUERI'

# Para cada campo canônico usado no cruzamento, os nomes normalizados (sem
# acentos, minúsculo) que podem aparecer como cabeçalho no Excel cadastral.
# A ordem da lista é a ordem de preferência (primeiro = mais específico).
_COL_ALIASES = {
    'Cód Epr':          ['cod emp', 'cod epr', 'codigo epr', 'codigo empresa', 'matricula'],
    'CPF':              ['cpf'],
    'RG':               ['rg', 'numero rg', 'no rg'],
    'UF RG':            ['uf rg', 'uf do rg', 'estado rg'],
    'Orgão RG':         ['orgao rg', 'orgao expedidor', 'orgao emissor', 'orgao exp'],
    'Data EX':          ['data ex', 'data emissao rg', 'data emissao', 'dt emissao', 'emissao rg'],
    'Data nascimento':  ['data nascimento', 'data de nascimento', 'dt nascimento', 'dt nasc', 'nascimento'],
    'Descrição cargo':  ['descricao cargo', 'desc cargo', 'cargo'],
    'Descrição Ccusto': ['descricao ccusto', 'desc ccusto', 'descricao centro de custo',
                         'centro de custo', 'ccusto'],
    'Descrição Dpto':   ['descricao dpto', 'desc dpto', 'descricao departamento',
                         'departamento', 'dpto'],
    'Endereço':         ['endereco', 'logradouro', 'end'],
    'Numero':           ['numero', 'num', 'numero end', 'numero endereco', 'nro'],
    'Complemento':      ['complemento', 'compl'],
    'Cep':              ['cep'],
    'Cidade':           ['cidade', 'municipio'],
    'UF End':           ['uf end', 'uf endereco', 'uf', 'estado end'],
    'Estado Civil':     ['estado civil'],
    'Nome Mae':         ['nome mae', 'nome da mae', 'mae', 'nome_mae'],
    'Administradora(Fornecedor)': [
        'administradora(fornecedor)', 'administradora (fornecedor)',
        'administradora', 'administradora do beneficio', 'fornecedor'
    ],
}

# Possiveis cabecalhos para planilha usada como fonte de extracao (no lugar do PDF).
_COL_ALIASES_FONTE = {
    'codigo':         ['codigo', 'cod emp', 'cod epr', 're', 'folha re', 'matricula', 'matricula vt'],
    'colaborador':    ['colaborador', 'nome', 'nome funcionario', 'nome do funcionario'],
    'periodo':        ['periodo', 'periodo trabalhado', 'periodo de dias trabalhados'],
    'quantidade':     ['quantidade', 'qtd', 'quantidade diaria', 'qtd diaria'],
    'valor_unitario': ['valor unitario', 'valor', 'valor diario', 'valor un'],
    'administradora': ['administradora', 'administradora do beneficio', 'operadora',
                       'administradora(fornecedor)', 'administradora (fornecedor)',
                       'rede recarga', 'beneficio do funcionario', 'beneficio', 'adm'],
}

# ──────────────────────────────────────────────────────────────────────────────
# Funções auxiliares de módulo
# ──────────────────────────────────────────────────────────────────────────────

def _norm(texto):
    """Remove acentos e converte para minúsculas para comparação robusta."""
    return unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode().lower().strip()


def _extrair_codigo(valor):
    """Extrai código numérico de célula, lidando com '123.0', '  123 ', etc."""
    if valor is None:
        return ''
    s = str(valor).strip().replace('\n', '').replace('\r', '')
    s = re.sub(r'\.0+$', '', s)
    m = re.match(r'^(\d+)', s)
    return m.group(1) if m else ''


def _pode_latin1(c):
    """Verifica se o caractere pode ser codificado em latin-1."""
    try:
        c.encode('latin-1')
        return True
    except UnicodeEncodeError:
        return False


def _normalizar_data_espacada(texto):
    return re.sub(r'\s+', '', texto)


def _limpar_nome_extraido(texto):
    texto = re.sub(r'\s+', ' ', texto).strip()
    # Remove dígitos misturados no meio do nome (artefato de pdfplumber em PDFs com
    # colunas adjacentes — ex: "SANTO1S" → "SANTOS", "OLIV9E1IR97A483" → "OLIVEIRA").
    # Só aplica se o texto contém letras (evita apagar matrículas puras).
    if re.search(r'[A-Za-zÀ-ÿ]', texto):
        texto = re.sub(r'\d', '', texto)
    # Normaliza espaços residuais e remove cauda vazia.
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto


# ──────────────────────────────────────────────────────────────────────────────
# Classe principal
# ──────────────────────────────────────────────────────────────────────────────

class ProcessadorVTCaixa:

    # ── Extração PDF ───────────────────────────────────────────────────

    def _extrair_pdf(self, pdf_path):
        rows = []
        avisos = []

        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                tabelas = pagina.extract_tables() or []
                if not tabelas:
                    tabela = pagina.extract_table()
                    if tabela:
                        tabelas = [tabela]
                if not tabelas:
                    continue

                for tabela in tabelas:
                    if not tabela:
                        continue

                    for linha in tabela:
                        if not linha or all(c is None for c in linha):
                            continue

                        codigo = _extrair_codigo(linha[0])
                        if not codigo:
                            continue

                        while len(linha) < 9:
                            linha.append(None)

                        def _cel(idx, ln=linha):
                            v = ln[idx]
                            return str(v).strip().replace('\n', ' ').replace('\r', '') if v is not None else ''

                        rows.append({
                            'codigo':         codigo,
                            'colaborador':    _limpar_nome_extraido(_cel(1)),
                            'periodo':        _cel(3),
                            'quantidade':     re.sub(r'\.0+$', '', _cel(5)),
                            'valor_unitario': _cel(6),
                            'administradora': _cel(8),
                        })

        if not rows:
            avisos.append('AVISO: Nenhuma linha tabular extraída do PDF; tentando leitura por texto.')
            rows = self._extrair_pdf_por_texto(pdf_path)

        if not rows:
            avisos.append(
                'AVISO: Nenhuma linha de dados extraída do PDF. '
                'Verifique se o PDF é o relatório correto e se o layout continua compatível com a extração.'
            )

        return rows, avisos

    def _extrair_pdf_por_texto(self, pdf_path):
        rows = []
        date_re = re.compile(r'(\d\s*\d\s*/\s*\d\s*\d\s*/\s*\d\s*\d\s*\d\s*\d)')

        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text() or ''
                if not texto:
                    continue

                iniciou_dados = False
                for linha in texto.splitlines():
                    linha = linha.strip()
                    if not linha:
                        continue

                    if 'Tipo Benefício:' in linha:
                        iniciou_dados = True
                        continue

                    if not iniciou_dados:
                        continue

                    if linha.isdigit():
                        continue

                    m_cod = re.match(r'^(\d+)\s+', linha)
                    if not m_cod:
                        continue

                    datas = list(date_re.finditer(linha))
                    if len(datas) < 3:
                        continue

                    codigo = m_cod.group(1)
                    prefixo = linha[m_cod.end():datas[0].start()].strip()
                    colaborador = _limpar_nome_extraido(prefixo)

                    periodo_inicio = _normalizar_data_espacada(datas[0].group(1))
                    periodo_fim = _normalizar_data_espacada(datas[1].group(1))
                    periodo = f'{periodo_inicio} a {periodo_fim}'

                    cauda = linha[datas[2].end():].strip()
                    m_tail = re.match(r'^(\d+)\s+(\d+,\d{2})\s+(\d+,\d{2})\s+(.+)$', cauda)
                    if not m_tail:
                        continue

                    rows.append({
                        'codigo':         codigo,
                        'colaborador':    colaborador,
                        'periodo':        periodo,
                        'quantidade':     m_tail.group(1),
                        'valor_unitario': m_tail.group(2),
                        'administradora': m_tail.group(4).strip(),
                    })

        return rows

    # ── Dias úteis ─────────────────────────────────────────────────────

    def _texto_celula(self, valor):
        if valor is None:
            return ''
        if isinstance(valor, datetime):
            return valor.strftime('%d/%m/%Y')
        if isinstance(valor, date):
            return valor.strftime('%d/%m/%Y')
        if isinstance(valor, float):
            if valor.is_integer():
                return str(int(valor))
            return str(valor)
        return str(valor).strip().replace('\n', ' ').replace('\r', '')

    def _iterar_abas_planilha_fonte(self, planilha_path):
        ext = os.path.splitext(planilha_path)[1].lower()

        if ext == '.xls':
            wb = xlrd.open_workbook(planilha_path)
            for idx in range(wb.nsheets):
                ws = wb.sheet_by_index(idx)
                linhas = []
                for row in range(ws.nrows):
                    linhas.append([ws.cell_value(row, col) for col in range(ws.ncols)])
                yield ws.name, linhas
            return

        wb = load_workbook(planilha_path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                linhas = [list(row) for row in ws.iter_rows(values_only=True)]
                yield ws.title, linhas
        finally:
            wb.close()

    def _extrair_fonte_planilha(self, planilha_path):
        avisos = []
        candidatos = []
        ordem_aba = 0

        for aba_nome, linhas in self._iterar_abas_planilha_fonte(planilha_path):
            ordem_aba += 1
            if not linhas:
                continue

            cabecalho = linhas[0]
            cabecalhos_norm = {}
            for i, val in enumerate(cabecalho):
                txt = self._texto_celula(val)
                if txt:
                    cabecalhos_norm[_norm(txt)] = i

            def _idx_chave(chave):
                for alias in _COL_ALIASES_FONTE[chave]:
                    if alias in cabecalhos_norm:
                        return cabecalhos_norm[alias]
                return None

            idx_codigo = _idx_chave('codigo')
            idx_nome = _idx_chave('colaborador')
            idx_periodo = _idx_chave('periodo')
            idx_qtd = _idx_chave('quantidade')
            idx_valor = _idx_chave('valor_unitario')
            idx_adm = _idx_chave('administradora')

            campos_fonte = [idx_nome, idx_periodo, idx_qtd, idx_valor, idx_adm]
            campos_encontrados = sum(1 for idx in campos_fonte if idx is not None)
            usou_cabecalho = idx_codigo is not None and campos_encontrados >= 2
            if usou_cabecalho:
                dados_linhas = linhas[1:]
            else:
                idx_codigo, idx_nome, idx_periodo = 0, 1, 3
                idx_qtd, idx_valor, idx_adm = 5, 6, 8
                dados_linhas = linhas

            def _get(ln, idx):
                if idx is None:
                    return ''
                if idx < 0 or idx >= len(ln):
                    return ''
                return self._texto_celula(ln[idx])

            rows = []
            for ln in dados_linhas:
                if not ln:
                    continue

                codigo = _extrair_codigo(_get(ln, idx_codigo))
                if not codigo:
                    continue

                rows.append({
                    'codigo':         codigo,
                    'colaborador':    _limpar_nome_extraido(_get(ln, idx_nome)),
                    'periodo':        _get(ln, idx_periodo),
                    'quantidade':     re.sub(r'\.0+$', '', _get(ln, idx_qtd)),
                    'valor_unitario': _get(ln, idx_valor),
                    'administradora': _get(ln, idx_adm),
                })

            if not rows:
                continue

            prioridade = 0 if usou_cabecalho else 1
            candidatos.append((prioridade, -len(rows), ordem_aba, aba_nome, rows, usou_cabecalho))

        if not candidatos:
            avisos.append('AVISO: Nenhuma linha valida foi extraida da planilha de fonte.')
            return [], avisos

        candidatos.sort(key=lambda item: (item[0], item[1], item[2]))
        _, _, _, aba_escolhida, rows_escolhidas, usou_cab = candidatos[0]
        modo = 'cabecalho' if usou_cab else 'posicao de colunas'
        avisos.append(f'Aba selecionada para extracao: {aba_escolhida} ({modo})')
        return rows_escolhidas, avisos

    def _extrair_pdf_fonte(self, pdf_path):
        """Extrai dados de PDF-fonte com cabeçalho (Setor, Re, Nome, ...)."""
        rows = []
        avisos = []

        with pdfplumber.open(pdf_path) as pdf:
            col_map = {}
            cabecalho_detectado = False

            for pagina in pdf.pages:
                tabelas = pagina.extract_tables() or []
                if not tabelas:
                    tabela = pagina.extract_table()
                    if tabela:
                        tabelas = [tabela]
                if not tabelas:
                    continue

                for tabela in tabelas:
                    if not tabela:
                        continue

                    for linha in tabela:
                        if not linha or all(c is None for c in linha):
                            continue

                        if not cabecalho_detectado:
                            cab_norm = {}
                            for j, cel in enumerate(linha):
                                if cel:
                                    cab_norm[_norm(str(cel))] = j
                            mapa = {}
                            for chave, aliases in _COL_ALIASES_FONTE.items():
                                for alias in aliases:
                                    if alias in cab_norm:
                                        mapa[chave] = cab_norm[alias]
                                        break
                            if mapa.get('codigo') is not None:
                                col_map = mapa
                                cabecalho_detectado = True
                            continue

                        def _cel(key, ln=linha):
                            idx = col_map.get(key)
                            if idx is None or idx >= len(ln):
                                return ''
                            v = ln[idx]
                            return str(v).strip().replace('\n', ' ').replace('\r', '') if v else ''

                        codigo = _extrair_codigo(_cel('codigo'))
                        if not codigo:
                            continue

                        rows.append({
                            'codigo':         codigo,
                            'colaborador':    _limpar_nome_extraido(_cel('colaborador')),
                            'periodo':        _cel('periodo'),
                            'quantidade':     re.sub(r'\.0+$', '', _cel('quantidade')),
                            'valor_unitario': _cel('valor_unitario'),
                            'administradora': _cel('administradora'),
                        })

        if not rows:
            avisos.append('AVISO: Nenhuma linha extraída do PDF-fonte com cabeçalho.')

        return rows, avisos

    def _extrair_fonte(self, fonte_path):
        ext = os.path.splitext(fonte_path)[1].lower()
        if ext == '.pdf':
            rows, avisos = self._extrair_pdf_fonte(fonte_path)
            if rows:
                return rows, avisos
            avisos.append('Fallback: tentando extração posicional do PDF.')
            rows_leg, avisos_leg = self._extrair_pdf(fonte_path)
            return rows_leg, avisos + avisos_leg
        if ext in ('.xls', '.xlsx', '.xlsm', '.xltx', '.xltm'):
            return self._extrair_fonte_planilha(fonte_path)
        raise ValueError('Formato de fonte nao suportado. Use PDF ou Excel (.xls/.xlsx).')

    def _calcular_dias_uteis(self, periodo_str):
        if not periodo_str:
            return 0

        partes = None
        s = periodo_str.strip()

        # Extrai as duas datas (dd/mm/yyyy) de qualquer forma que apareçam.
        # Aceita: 'a', 'A', '-', 'até', espaço duplo, ou qualquer separador entre elas.
        datas = re.findall(r'\d{2}/\d{2}/\d{4}', s)
        if len(datas) >= 2:
            partes = [datas[0], datas[1]]

        # Fallback: dd-mm-yyyy ou dd.mm.yyyy
        if not partes:
            datas_alt = re.findall(r'\d{2}[-\\.]\d{2}[-\\.]\d{4}', s)
            if len(datas_alt) >= 2:
                partes = [
                    re.sub(r'[-\\.]', '/', datas_alt[0]),
                    re.sub(r'[-\\.]', '/', datas_alt[1]),
                ]

        if not partes:
            return 0

        try:
            inicio = datetime.strptime(partes[0], '%d/%m/%Y').date()
            fim    = datetime.strptime(partes[1], '%d/%m/%Y').date()
        except ValueError:
            return 0

        dias = 0
        atual = inicio
        while atual <= fim:
            if atual.weekday() < 5:
                dias += 1
            atual = date.fromordinal(atual.toordinal() + 1)
        return dias

    # ── Formatadores ───────────────────────────────────────────────────

    def _formatar_cpf(self, valor):
        if valor is None or valor == '':
            return ''
        if isinstance(valor, float):
            valor = str(int(valor))
        # Extrai apenas dígitos (remove pontuação '123.456.789-00', espaços, etc.)
        digits = re.sub(r'\D', '', str(valor))
        if not digits:
            return ''
        # CPF tem 11 dígitos; zeros à esquerda são preservados.
        return digits.zfill(11)

    def _formatar_rg(self, valor):
        if valor is None or valor == '':
            return ''
        if isinstance(valor, float):
            return str(int(valor))
        s = str(valor).strip()
        s = re.sub(r'\.0+$', '', s)
        digits = re.sub(r'\D', '', s)
        return digits if digits else s

    def _formatar_data(self, valor, wb):
        if valor is None or valor == '':
            return ''
        if isinstance(valor, float):
            try:
                dt = xlrd.xldate_as_datetime(valor, wb.datemode)
                return dt.strftime('%d/%m/%Y')
            except Exception:
                return ''
        s = str(valor).strip()
        # Já no formato dd/mm/yyyy → devolve direto
        if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
            return s
        # ISO yyyy-mm-dd ou yyyy-mm-dd hh:mm:ss
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            return f'{m.group(3)}/{m.group(2)}/{m.group(1)}'
        return s

    def _formatar_cep(self, valor):
        if valor is None or valor == '':
            return ''
        if isinstance(valor, float):
            valor = str(int(valor))
        digits = re.sub(r'\D', '', str(valor))
        if not digits:
            return ''
        # CEP tem 8 dígitos; zeros à esquerda são preservados.
        return digits.zfill(8)[:8]

    def _formatar_numero(self, valor):
        if valor is None or valor == '':
            return ''
        if isinstance(valor, float):
            return str(int(valor))
        s = str(valor).strip()
        # Remove '.0' final se veio de float→str, mas preserva '123-A', '123/B'.
        return re.sub(r'\.0+$', '', s)

    # ── Carregar Excel ─────────────────────────────────────────────────

    def _carregar_excel(self, xls_path):
        wb = _abrir_workbook_cadastral(xls_path)
        avisos = []

        ws = None
        for idx in range(wb.nsheets):
            candidato = wb.sheet_by_index(idx)
            cabecalhos = {
                _norm(candidato.cell_value(0, col))
                for col in range(candidato.ncols)
                if candidato.cell_value(0, col)
            }
            if 'cod emp' in cabecalhos or 'cod epr' in cabecalhos:
                ws = candidato
                avisos.append(f'Aba selecionada no Excel: {candidato.name}')
                break

        if ws is None:
            ws = wb.sheet_by_index(0)
            avisos.append(
                f"Aviso: aba com 'Cód Emp'/'Cód Epr' não encontrada; usando a primeira aba: {ws.name}"
            )

        cabecalhos_norm = {}
        cabecalhos_orig = {}
        for col in range(ws.ncols):
            val = ws.cell_value(0, col)
            if val:
                n = _norm(val)
                cabecalhos_norm[n] = col
                cabecalhos_orig[n] = str(val).strip()

        avisos.append(f'Colunas no Excel: {list(cabecalhos_orig.values())}')

        def _idx(nome_canonico):
            """Localiza coluna no Excel: exato → aliases conhecidos. Sem fuzzy."""
            alvo = _norm(nome_canonico)
            # 1) Match exato pelo nome canônico
            if alvo in cabecalhos_norm:
                return cabecalhos_norm[alvo]
            # 2) Tabela de aliases: testa cada variante conhecida, em ordem.
            for alias in _COL_ALIASES.get(nome_canonico, []):
                if alias in cabecalhos_norm:
                    return cabecalhos_norm[alias]
            return None

        mapa_colunas = {
            'Cód Epr':          _idx('Cód Epr'),
            'CPF':              _idx('CPF'),
            'RG':               _idx('RG'),
            'UF RG':            _idx('UF RG'),
            'Orgão RG':         _idx('Orgão RG'),
            'Data nascimento':  _idx('Data nascimento'),
            'Descrição cargo':  _idx('Descrição cargo'),
            'Descrição Ccusto': _idx('Descrição Ccusto'),
            'Descrição Dpto':   _idx('Descrição Dpto'),
            'Endereço':         _idx('Endereço'),
            'Numero':           _idx('Numero'),
            'Complemento':      _idx('Complemento'),
            'Cep':              _idx('Cep'),
            'Cidade':           _idx('Cidade'),
            'UF End':           _idx('UF End'),
            'Estado Civil':     _idx('Estado Civil'),
            'Data EX':          _idx('Data EX'),
            'Nome Mae':         _idx('Nome Mae'),
            'Administradora(Fornecedor)': _idx('Administradora(Fornecedor)'),
        }
        idx_cod     = mapa_colunas['Cód Epr']
        idx_cpf     = mapa_colunas['CPF']
        idx_rg      = mapa_colunas['RG']
        idx_uf_rg   = mapa_colunas['UF RG']
        idx_org_rg  = mapa_colunas['Orgão RG']
        idx_dt_nasc = mapa_colunas['Data nascimento']
        idx_cargo   = mapa_colunas['Descrição cargo']
        idx_ccusto  = mapa_colunas['Descrição Ccusto']
        idx_dpto    = mapa_colunas['Descrição Dpto']
        idx_end     = mapa_colunas['Endereço']
        idx_num     = mapa_colunas['Numero']
        idx_comp    = mapa_colunas['Complemento']
        idx_cep     = mapa_colunas['Cep']
        idx_cidade  = mapa_colunas['Cidade']
        idx_uf_end  = mapa_colunas['UF End']
        idx_est_civ = mapa_colunas['Estado Civil']
        idx_dt_ex   = mapa_colunas['Data EX']
        idx_nome_mae = mapa_colunas['Nome Mae']
        idx_adm_forn = mapa_colunas['Administradora(Fornecedor)']

        if idx_cod is None:
            raise ValueError(
                f"Coluna 'Cód Epr' não encontrada.\n"
                f"Disponíveis: {list(cabecalhos_orig.values())}"
            )

        faltantes = [nome for nome, idx in mapa_colunas.items() if idx is None]
        if faltantes:
            avisos.append(
                f'AVISO: colunas não encontradas no Excel (campos ficarão vazios): {faltantes}'
            )

        avisos.append(
            f'Mapeamento: cod={idx_cod}, cpf={idx_cpf}, rg={idx_rg}, '
            f'nascimento={idx_dt_nasc}, cargo={idx_cargo}, ccusto={idx_ccusto}, '
            f'endereço={idx_end}, numero={idx_num}, cep={idx_cep}'
        )

        dados = {}
        for row in range(1, ws.nrows):
            cod_raw = ws.cell_value(row, idx_cod)
            if cod_raw is None or cod_raw == '':
                continue
            chave = _extrair_codigo(cod_raw) or str(cod_raw).strip()
            if not chave:
                continue

            def _val(idx, r=row):
                if idx is None:
                    return ''
                v = ws.cell_value(r, idx)
                if v is None or v == '':
                    return ''
                if isinstance(v, float):
                    # Inteiros guardados como float no xlrd (ex: 1234.0) → "1234"
                    if v.is_integer():
                        return str(int(v))
                    return str(v)
                return str(v).strip()

            dados[chave] = {
                'CPF':              self._formatar_cpf(ws.cell_value(row, idx_cpf) if idx_cpf is not None else None),
                'RG':               self._formatar_rg(ws.cell_value(row, idx_rg) if idx_rg is not None else None),
                'UF RG':            _val(idx_uf_rg),
                'Orgão RG':         _val(idx_org_rg),
                'Data nascimento':  self._formatar_data(ws.cell_value(row, idx_dt_nasc) if idx_dt_nasc is not None else None, wb),
                'Descrição cargo':  _val(idx_cargo),
                'Descrição Ccusto': _val(idx_ccusto),
                'Descrição Dpto':   _val(idx_dpto),
                'Endereço':         _val(idx_end),
                'Numero':           self._formatar_numero(ws.cell_value(row, idx_num) if idx_num is not None else None),
                'Complemento':      _val(idx_comp),
                'Cep':              self._formatar_cep(ws.cell_value(row, idx_cep) if idx_cep is not None else None),
                'Cidade':           _val(idx_cidade),
                'UF End':           _val(idx_uf_end),
                'Estado Civil':     _val(idx_est_civ),
                'Data EX':          self._formatar_data(ws.cell_value(row, idx_dt_ex) if idx_dt_ex is not None else None, wb),
                'Nome Mae':         _val(idx_nome_mae),
                'Administradora(Fornecedor)': _val(idx_adm_forn),
            }

        if not dados:
            avisos.append(
                'AVISO: Nenhum funcionário carregado do Excel cadastral. '
                'Verifique se a aba selecionada contém cabeçalho e dados abaixo da coluna Cód Epr.'
            )

        return dados, avisos

    # ── Cruzamento ─────────────────────────────────────────────────────

    def _limpar_valor_unitario(self, valor_str):
        """Converte valor monetario para formato com virgula decimal (ex: '1234,56')."""
        if valor_str is None:
            return ''
        v = str(valor_str).replace('R$', '').replace('\xa0', ' ').strip()
        if not v:
            return ''
        # Formato pt-BR: "1.234,56" → remove separador de milhar, troca decimal.
        if ',' in v:
            v = v.replace('.', '').replace(',', '.')
        # Sem vírgula: pode já estar em US ("1234.56") — preservar.
        try:
            return f'{float(v):.2f}'.replace('.', ',')
        except ValueError:
            return ''

    def _sanitizar(self, texto):
        """Remove quebras de linha, tabs e normaliza espaços. Nunca retorna None."""
        if texto is None:
            return ''
        s = str(texto).replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
        return re.sub(r'\s+', ' ', s).strip()

    # Tabela: (substring operadora uppercase, valor unitário normalizado) → código benefício
    # Valor None = qualquer valor (ex: VIANOVA TOP independe do valor)
    _CODIGOS_BENEFICIO = [
        ('SPTRANS', '11,64', '701'),
        ('SPTRANS', '11,84', '695'),
        ('SPTRANS', '22,64', '698'),
        ('VIANOVA - TOP', None, '13665'),
        ('ONPAG RADIAL', None, '13671'),
        ('GUARUPASS', None, '9100'),
        ('ITAQUA', None, '8498'),
        ('CITY - BAIXADA', None, '10946'),
        ('SIM MAUA', None, '36826'),
        ('BR MOBILIDADEBAIXADA SANTISTA', None, '38237'),
        ('MORATENSE', None, '7701'),
        ('MOGI PASSE', None, '8699'),
        ('BEM FRANCO DA ROCHA', None, '15107'),
        ('BEM OSASCO', None, '15102'),
        ('PIRACICABANA', None, '48116'),
        ('MARILIA CARD', None, '48258'),
        ('PIRA MOBILIDADE', None, '48533'),
        ('RÁPIDO CAMPINAS SOROCABA', None, '37824'),
        ('SOU SÃO SEBASTIÃO', None, '13378'),
        ('TRANSURC', None, '1310'),
        ('URBES - SOROCABA', None, '1910'),
        ('BILHETE SÃO JOSEENSE', None, '1265'),
        ('VIAÇÃO JACAREÍ', None, '1494'),
        ('SOU ITANHAEM', None, '13002'),
        ('VALE SUL', None, '6278'),
        ('VIAÇÃO NA MONTANHA', None, '37257'),
        ('CARTAO BÁ - BAGÉ', None, '4775'),
        ('CARTAO RIO GRANDE', None, '14067'),
        ('CAXIAS URBANO', None, '3268'),
        ('COLEO - BEM', None, '15196'),
        ('COLEURB', None, '14736'),
        ('PRATI PELOTAS', None, '38358'),
        ('SIM SANTA MARIA-ATU', None, '11962'),
        ('TEUTO - STADBUS', None, '11818'),
        ('TRI', None, '6084'),
        ('URBANO ERECHIM', None, '11120'),
        ('VINO', None, '9044'),
        ('SOU UBATUBA', None, '1690'),
        ('USEPASSE', None, '37896'),
        ('VIMSA', None, '4910'),
        ('MIXMOB', None, '14587'),
        ('TCS', None, '9558'),
        ('CIDADAO MANAUS', None, '4481'),
        ('VIACAO BOA VISTA', None, '3275'),
        ('VIAÇÃO BOA VISTA', None, '3275'),
        ('RIOCARD', None, '2232'),
        ('JAÉ', None, '14816'),
        ('JAE', None, '14816'),
        ('BRB MOBILIDADE', None, '2839'),
        ('CARTAO UTB', None, '10568'),
        ('AMAZONIA INTER', None, '14039'),
        ('TAGUATUR AGUAS', None, '10985'),
        ('TRANSURB BAURU', None, '1858'),
        ('ANDREENSE', None, '9022'),
        ('URBAN', None, '1473'),
        ('SIND CARD', None, '5696'),
        ('CARTAO LEGAL', None, '9702'),
        ('DEL REY TRANSPORTES', None, '6308'),
        ('EMBU', None, '15406'),
        ('SOU CUBATÃO', None, '11363'),
        ('SOU DIADEMA', None, '9780'),
        ('BUS FACIL RIBEIRAO PIRES', None, '15086'),
    ]

    # Substituições de departamento (original -> substituto) aplicadas no processamento.
    _DEPART_MAP = {
        'CEF LESTE 10 SP 4719/2022': 'CEF 10 84',
        'CEF 17 CONTRATO 477/2026':  'CEF 17 LIMPEZA',
        'CEF 12 AMAZONAS - AM e RR': 'CEF 12 87',
        'CEF BAIXADA 11 SP 4820/2022': 'CEF 11 85',
        'POLICIA FED SHOP FLAMINGO':      'PF SHOPPING FLAMINGO',
        'DPF - SUPERINT POLICIA FEDERAL': 'PF SAO PAULO',
        'SP - DPF CAMPINAS':              'PF CAMPINAS',
        'SP - DPF SANTOS':                'PF SANTOS',
        'SP - DPF SAO JOSE DOS CAMPOS':   'PF SAO JOSE DOS CAMPOS',
        'SP - DPF SOROCABA':              'PF SOROCABA',
        'POLIC FED AEROPORTO GUARULHOS':  'PF GUARULHOS',
        'SP - DPF MARILIA':               'PF MARILIA',
        'SP - DPF RIBEIRAO PRETO':        'PF RIBEIRAO PRETO',
        'SP - DPF PIRACICABA':            'PF PIRACICABA',
        'SP - DPF BAURU':                 'PF BAURU',
        'SP - DPF COMPLEXO AGUA BRANCA':  'PF AGUA BRANCA',
        'POLICIA FED CONGONHAS':          'PF CONGONHAS',
        'B BRASIL RJ 2022.7421.6922':     'BB RJ 89',
        'CEF 14 DF':                      'CEF 14 DF 90',
        'CEF 15 RS 4916':                 'CEF 15 RS',
        'ELETRONUCLEAR RECEP 4500070400': 'ELETRONUCLEAR RECEP 97',
    }

    def _resolver_codigo_beneficio(self, administradora, valor_unitario, extras=None):
        """Retorna o código de benefício quando a operadora+valor bate uma regra, ou None.

        `extras` (regras personalizadas, mesmo formato de _CODIGOS_BENEFICIO)
        é consultado antes das regras embutidas.
        """
        adm_up = administradora.upper()
        for operadora, valor_regra, codigo in list(extras or []) + self._CODIGOS_BENEFICIO:
            if operadora not in adm_up:
                continue
            if valor_regra is None or valor_unitario == valor_regra:
                return codigo
        return None

    def _cruzar_dados(self, pdf_rows, excel_data, codigos_extras=None, depart_extras=None):
        registros = []
        nao_encontrados = []

        for linha in pdf_rows:
            codigo = linha['codigo']
            ex = excel_data.get(codigo)

            if ex is None:
                nao_encontrados.append(f"{codigo} - {linha['colaborador']}")
                continue

            registros.append({
                'CNPJ':                        EMPRESA_CNPJ,
                'CEP':                         EMPRESA_CEP,
                'LOGRADOURO':                  EMPRESA_LOGRADOURO,
                'NÚMERO':                      EMPRESA_NUMERO,
                'COMPLEMENTO':                 EMPRESA_COMPLEMENTO,
                'PONTO REFERENCIA':            '',
                'UF':                          EMPRESA_UF,
                'ESTADO':                      EMPRESA_ESTADO,
                'MATRÍCULA':                   codigo,
                'NOME DO FUNCIONÁRIO':         self._sanitizar(linha['colaborador']),
                'CPF':                         ex['CPF'],
                'RG':                          ex['RG'] or ex['CPF'],
                'DATA DE NASCIMENTO':          ex['Data nascimento'],
                'CARGO':                       self._sanitizar(ex['Descrição cargo']),
                'DEPARTAMENTO':                self._sanitizar(
                    ex['Descrição Dpto']
                    if ex.get('Descrição Ccusto', '').strip().upper() == 'DEP POLICIA FEDERAL SP'
                    else ex['Descrição Ccusto']
                ),
                'NOME DA MÃE':                 self._sanitizar(ex['Nome Mae']),
                'BENEFÍCIO DO FUNCIONÁRIO':    self._sanitizar(linha['administradora']),
                'VALOR UNITÁRIO':              self._limpar_valor_unitario(linha['valor_unitario']),
                'QUANTIDADE DIÁRIA':           '2',
                'PERÍODO DE DIAS TRABALHADOS': linha.get('quantidade', ''),
                'TIPO VALOR':                  '',
                'REDE RECARGA':                '3' if 'RIOCARD' in self._sanitizar(linha['administradora']).upper() else '',
                'CEP RESIDENCIAL':             ex['Cep'],
                'LOGRADOURO RESIDENCIAL':      self._sanitizar(ex['Endereço']),
                'NÚMERO RESIDENCIAL':          ex['Numero'],
                'COMPLEMENTO RESIDENCIAL':     self._sanitizar(ex['Complemento']),
                'ESTADO CIVIL':                self._sanitizar(ex['Estado Civil']),
                'DATA DE EMISSÃO DO RG':       ex['Data EX'],
                'ÓRGÃO EXPEDIDOR':             self._sanitizar(ex['Orgão RG']),
                'ESTADO EMISSÃO RG':           self._sanitizar(ex['UF RG']),
            })

        for reg in registros:
            chave_mat = next((k for k in reg.keys() if 'MATR' in k), None)
            chave_benef = next((k for k in reg.keys() if 'BENEF' in k and 'FUNCION' in k), None)
            if not chave_mat or not chave_benef:
                continue
            ex = excel_data.get(str(reg[chave_mat]).strip())
            if not ex:
                continue
            adm_forn = ex.get('Administradora(Fornecedor)')
            if adm_forn:
                reg[chave_benef] = self._sanitizar(adm_forn)

        # Substituições de departamento (personalizadas sobrepõem o mapa embutido)
        depart_map = {**self._DEPART_MAP, **(depart_extras or {})}
        for reg in registros:
            depart = reg.get('DEPARTAMENTO', '')
            if depart in depart_map:
                reg['DEPARTAMENTO'] = depart_map[depart]

        # Substituir BENEFÍCIO DO FUNCIONÁRIO pelo código quando operadora+valor bater uma regra
        for reg in registros:
            chave_benef = next((k for k in reg.keys() if 'BENEF' in k and 'FUNCION' in k), None)
            if not chave_benef:
                continue
            codigo = self._resolver_codigo_beneficio(
                reg[chave_benef],
                reg.get('VALOR UNITÁRIO', ''),
                codigos_extras,
            )
            if codigo:
                reg[chave_benef] = codigo

        return registros, nao_encontrados

    # ── Geração CSV ────────────────────────────────────────────────────

    def _gerar_csv(self, registros, output_path):
        """Grava CSV em latin-1 (;), CRLF. Retorna avisos sobre chars substituídos."""
        avisos = []
        codec = 'latin-1'

        # 1) Sanity check dos nomes de campo — todos presentes e strings?
        for reg in registros:
            mat = reg.get('MATRÍCULA', '?')
            for col in COLUNAS_CSV:
                if col not in reg:
                    reg[col] = ''
                    continue
                if reg[col] is None:
                    reg[col] = ''
                elif not isinstance(reg[col], str):
                    reg[col] = str(reg[col])

            # 2) Detecta caracteres não-latin1 antes de gravar (ficariam '?').
            for col in COLUNAS_CSV:
                val = reg[col]
                problemas = sorted({c for c in val if not _pode_latin1(c)})
                if problemas:
                    avisos.append(
                        f"Matrícula {mat} / {col}: "
                        f"char(s) fora do latin-1 substituído(s) por '?': {problemas}"
                    )

        with open(output_path, 'w', newline='', encoding=codec, errors='replace') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=COLUNAS_CSV,
                delimiter=';',
                quotechar='"',
                quoting=csv.QUOTE_MINIMAL,
                lineterminator='\r\n',
                extrasaction='ignore',
            )
            writer.writeheader()
            writer.writerows(registros)

        return avisos

    # ── Orquestrador ───────────────────────────────────────────────────

    def processar(self, fonte_path, xls_path, output_path,
                  progress_cb=None, codigos_extras=None, depart_extras=None):
        def _prog(pct, msg):
            if progress_cb:
                progress_cb(pct, msg)

        ext_fonte = os.path.splitext(fonte_path)[1].lower()
        tipo_fonte = 'PDF' if ext_fonte == '.pdf' else 'Excel'

        _prog(5, 'Lendo fonte de extracao...')
        pdf_rows, avisos_pdf = self._extrair_fonte(fonte_path)
        _prog(35, f'Fonte ({tipo_fonte}): {len(pdf_rows)} linha(s) encontrada(s).')
        for av in avisos_pdf:
            _prog(35, av)
        if pdf_rows:
            _prog(35, f'1º código PDF: {pdf_rows[0]["codigo"]}  |  último: {pdf_rows[-1]["codigo"]}')

        _prog(40, 'Carregando Excel cadastral...')
        excel_data, avisos_xls = self._carregar_excel(xls_path)
        _prog(60, f'Excel: {len(excel_data)} funcionário(s).')
        for av in avisos_xls:
            _prog(60, av)
        if excel_data:
            _prog(60, f'Primeiros códigos Excel: {list(excel_data.keys())[:3]}')

        if not pdf_rows:
            raise ValueError(
                'Nenhuma linha válida foi extraída do PDF VT Caixa. '
                'Verifique se o PDF informado é o relatório correto e se o layout continua compatível com a extração.'
            )

        if not excel_data:
            raise ValueError(
                'Nenhum cadastro válido foi carregado do Excel. '
                'Verifique a aba usada e se a coluna Cód Epr possui dados.'
            )

        _prog(65, 'Cruzando dados...')
        registros, nao_encontrados = self._cruzar_dados(
            pdf_rows, excel_data, codigos_extras, depart_extras)
        _prog(80, f'{len(registros)} correspondência(s) | {len(nao_encontrados)} sem correspondência.')

        if pdf_rows and not registros:
            raise ValueError(
                'Nenhuma correspondência encontrada entre o PDF e o Excel cadastral. '
                'Verifique se o cadastro está na aba correta e se a coluna Cód Epr corresponde aos códigos do PDF.'
            )

        _prog(85, 'Gerando CSV...')
        avisos_csv = self._gerar_csv(registros, output_path)
        for av in avisos_csv:
            _prog(87, f'AVISO encoding: {av}')
        _prog(90, f'CSV salvo em {output_path}')

        _prog(100, 'Concluído!')

        return {
            'total_pdf':       len(pdf_rows),
            'total_fonte':     len(pdf_rows),
            'tipo_fonte':      tipo_fonte,
            'total_ok':        len(registros),
            'nao_encontrados': nao_encontrados,
            'avisos_csv':      avisos_csv,
        }
