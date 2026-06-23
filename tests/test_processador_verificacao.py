# tests/test_processador_verificacao.py
import pytest
from processador import ProcessadorOcorrencias

proc = ProcessadorOcorrencias()

class FakePage:
    """Simula uma página do pdfplumber com extract_text()."""
    def __init__(self, text):
        self._text = text
    def extract_text(self):
        return self._text

class FakePDF:
    def __init__(self, pages):
        self.pages = pages
    def __enter__(self): return self
    def __exit__(self, *a): pass


def test_extrair_ocorrencias_texto_basico(monkeypatch):
    texto = (
        "SILVA JOAO          12345  ... AT AT FA\n"
        "SOUZA MARIA         67890  ... AT AT AT\n"
    )
    import pdfplumber

    class FakePDF2:
        pages = [FakePage(texto)]
        def __enter__(self): return self
        def __exit__(self, *a): pass

    monkeypatch.setattr(pdfplumber, 'open', lambda path: FakePDF2())
    resultado = proc.extrair_ocorrencias_texto('fake.pdf', ['AT', 'FA'])

    assert '12345' in resultado
    assert resultado['12345']['ocorrencias']['AT'] == 2
    assert resultado['12345']['ocorrencias']['FA'] == 1
    assert '67890' in resultado
    assert resultado['67890']['ocorrencias']['AT'] == 3


def test_extrair_ocorrencias_texto_sem_ocorrencias(monkeypatch):
    import pdfplumber

    class FakePDF2:
        pages = [FakePage("SILVA JOAO  12345  SD SD\n")]
        def __enter__(self): return self
        def __exit__(self, *a): pass

    monkeypatch.setattr(pdfplumber, 'open', lambda path: FakePDF2())
    resultado = proc.extrair_ocorrencias_texto('fake.pdf', ['AT'])
    assert resultado == {}


def test_reconciliar_sem_conflito():
    v1 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 2, 'FA': 1}}}
    v2 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 2, 'FA': 1}}}
    resultado = proc.reconciliar([v1, v2], ['AT', 'FA'])
    assert '12345' in resultado['concordantes']
    assert resultado['conflitos'] == []


def test_reconciliar_com_conflito():
    v1 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 2}}}
    v2 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 1}}}
    resultado = proc.reconciliar([v1, v2], ['AT'])
    assert resultado['concordantes'] == {}
    assert len(resultado['conflitos']) == 1
    c = resultado['conflitos'][0]
    assert c['re'] == '12345'
    assert c['codigo'] == 'AT'
    assert c['valores']['v1'] == 2
    assert c['valores']['v2'] == 1
    assert c['sugestao'] == 2  # v1 e v2 empatam → usa o maior


def test_reconciliar_maioria_vence():
    v1 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 2}}}
    v2 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 1}}}
    ia = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 2}}}
    resultado = proc.reconciliar([v1, v2, ia], ['AT'])
    # 2 votos em AT=2, 1 voto em AT=1 → sem conflito (maioria clara)
    assert '12345' in resultado['concordantes']
    assert resultado['concordantes']['12345']['ocorrencias']['AT'] == 2
    assert resultado['conflitos'] == []


def test_reconciliar_preserva_codigos_concordantes_de_re_com_conflito():
    # FA concorda nas duas varreduras; AT conflita. O FA não pode ser perdido
    # quando o RE vai para o diálogo de conflitos.
    v1 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'FA': 2, 'AT': 2}}}
    v2 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'FA': 2, 'AT': 1}}}
    resultado = proc.reconciliar([v1, v2], ['AT', 'FA'])
    assert len(resultado['conflitos']) == 1
    assert resultado['conflitos'][0]['codigo'] == 'AT'
    assert resultado['concordantes']['12345']['ocorrencias'] == {'FA': 2}


def test_reconciliar_re_ausente_em_uma_camada():
    # RE presente na V1 mas não na V2 → conflito com v2=0
    v1 = {'12345': {'nome': 'SILVA', 'ocorrencias': {'AT': 1}}}
    v2 = {}
    resultado = proc.reconciliar([v1, v2], ['AT'])
    assert len(resultado['conflitos']) == 1
    c = resultado['conflitos'][0]
    assert c['valores']['v1'] == 1
    assert c['valores']['v2'] == 0


def test_verificar_com_ia_retorna_none_sem_api_key():
    resultado = proc.verificar_com_ia('fake.pdf', ['AT', 'FA'], api_key='', modelo='gemini-1.5-flash')
    assert resultado is None


def test_verificar_com_ia_retorna_none_em_erro(monkeypatch):
    import pypdfium2 as pdfium

    def raise_err(path):
        raise Exception("pdf error")

    monkeypatch.setattr(pdfium, 'PdfDocument', raise_err)
    resultado = proc.verificar_com_ia('fake.pdf', ['AT'], api_key='fake-key', modelo='gemini-1.5-flash')
    assert resultado is None


class _FakeBitmap:
    def to_pil(self):
        from PIL import Image
        return Image.new('RGB', (10, 10))


class _FakePdfiumPage:
    def render(self, scale):
        return _FakeBitmap()


class _FakePdfiumDoc:
    def __len__(self): return 1
    def __getitem__(self, i): return _FakePdfiumPage()


def _fake_genai_client(resposta_texto):
    """Fake da API nova google.genai: Client(api_key=...).models.generate_content(...)."""
    class FakeResponse:
        text = resposta_texto

    class FakeModels:
        def generate_content(self, model, contents):
            return FakeResponse()

    class FakeClient:
        def __init__(self, api_key=None):
            self.models = FakeModels()

    return FakeClient


def test_verificar_com_ia_parseia_json_valido(monkeypatch):
    import pypdfium2 as pdfium
    import google.genai as genai

    resposta_json = '[{"re": "12345", "nome": "SILVA", "ocorrencias": {"AT": 2, "FA": 1}}]'
    monkeypatch.setattr(pdfium, 'PdfDocument', lambda path: _FakePdfiumDoc())
    monkeypatch.setattr(genai, 'Client', _fake_genai_client(resposta_json))

    resultado = proc.verificar_com_ia('fake.pdf', ['AT', 'FA'], api_key='fake-key', modelo='gemini-1.5-flash')
    assert resultado is not None
    assert '12345' in resultado
    assert resultado['12345']['ocorrencias']['AT'] == 2


def test_processar_retorna_e_grava_nao_encontrados(tmp_path):
    """RE presente no PDF mas ausente da planilha vira 'não localizado',
    é retornado e gravado na aba 'Não localizados' da saída."""
    from openpyxl import Workbook, load_workbook

    xlsx = tmp_path / "pedido.xlsx"
    out = tmp_path / "pedido_out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Folha RE", "MOTIVO"])
    ws.append(["12345", ""])  # só este RE existe na planilha
    wb.save(xlsx)

    dados = {
        "12345": {"nome": "SILVA", "ocorrencias": {"AT": 2}},   # bate
        "99999": {"nome": "AUSENTE", "ocorrencias": {"AT": 1}},  # não bate
    }
    res = proc.processar(str(xlsx), str(xlsx), str(out), ["AT"],
                         dados_externos=dados)

    nomes = {x["re"]: x for x in res["nao_encontrados"]}
    assert "99999" in nomes
    assert "12345" not in nomes
    assert nomes["99999"]["nome"] == "AUSENTE"

    wb_out = load_workbook(out)
    assert "Não localizados" in wb_out.sheetnames
    linhas = list(wb_out["Não localizados"].iter_rows(values_only=True))
    assert linhas[0] == ("Folha RE", "Nome", "Motivo")
    assert any(r[0] == "99999" for r in linhas[1:])


def test_verificar_com_ia_remove_cerca_markdown(monkeypatch):
    import pypdfium2 as pdfium
    import google.genai as genai

    resposta = ('```json\n'
                '[{"re": "12345", "nome": "SILVA", "ocorrencias": {"AT": 2}}]\n'
                '```')
    monkeypatch.setattr(pdfium, 'PdfDocument', lambda path: _FakePdfiumDoc())
    monkeypatch.setattr(genai, 'Client', _fake_genai_client(resposta))

    resultado = proc.verificar_com_ia('fake.pdf', ['AT'], api_key='fake-key', modelo='gemini-1.5-flash')
    assert resultado is not None
    assert resultado['12345']['ocorrencias']['AT'] == 2
